from __future__ import annotations

# ruff: noqa: E402
import hashlib
import json
import os
import re
import sys
import tempfile
import unicodedata
import zipfile
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPOSITORY_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPOSITORY_ROOT / "services/api/src"))

from flow_api.data_contract.contract import load_contract  # type: ignore[import-untyped]
from flow_api.data_contract.workbook import workbook_rows  # type: ignore[import-untyped]
from flow_api.fixtures.generator import build_reference_package  # type: ignore[import-untyped]

CONTRACT_PATH = REPOSITORY_ROOT / "templates/excel/flow_v1_contract.yaml"
WORKBOOK_PATH = REPOSITORY_ROOT / "fixtures/workbooks/external_logistics_nonstandard_v1.xlsx"
MANIFEST_PATH = REPOSITORY_ROOT / "fixtures/intake/nonstandard_manifest.json"
FIXED_TIME = datetime(2026, 8, 30, 0, 0, 0)
ZIP_TIME = (2026, 8, 30, 0, 0, 0)

SHEET_ORDER = (
    ("客户档案", "customer_master"),
    ("业务明细", "operating_actual"),
    ("管理账", "financial_actual"),
    ("应收账龄", "ar_collection"),
    ("月度计划", "monthly_budget"),
    ("物流服务", "logistics_product"),
    ("组织区域", "organization_region"),
    ("管理科目表", "management_account"),
    ("批次信息", "analysis_batch"),
)

FIELD_ALIASES = {
    "batch_code": "分析批号",
    "contract_version": "目标数据版本",
    "analysis_start_month": "分析起月",
    "analysis_end_month": "分析止月",
    "comparison_start_month": "同比起月",
    "comparison_end_month": "同比止月",
    "currency": "本位币",
    "actual_scenario_code": "实际情景",
    "budget_scenario_code": "计划情景",
    "budget_version_label": "计划版本",
    "generated_at": "生成时间",
    "record_id": "来源记录号",
    "month_key": "业务月份",
    "organization_code": "分公司编码",
    "customer_code": "客户编号",
    "logistics_product_code": "服务产品编码",
    "region_code": "经营区域编码",
    "order_count": "订单票数",
    "shipment_count": "履约件数",
    "revenue": "营业收入(元)",
    "warehousing_cost": "仓储成本(元)",
    "transportation_cost": "运输成本(元)",
    "other_direct_cost": "其他直接成本(元)",
    "management_account_code": "管理口径科目",
    "amount": "发生额(元)",
    "customer_segment_code": "客群编号",
    "scenario_code": "情景编号",
    "metric_code": "预算指标",
    "invoice_number": "发票号码",
    "aging_bucket": "账龄段",
    "receivable_balance": "应收余额(元)",
    "due_amount": "到期金额(元)",
    "overdue_amount": "逾期金额(元)",
    "collected_amount": "本期回款(元)",
    "customer_name": "客户全称",
    "industry": "所属行业",
    "tier": "客户级别",
    "credit_term_days": "信用期(天)",
    "customer_segment_name": "客群名称",
    "logistics_product_name": "物流服务名称",
    "level": "层次",
    "parent_code": "上级编号",
    "entity_type": "档案类型",
    "entity_code": "档案编码",
    "entity_name": "档案名称",
    "province": "省份",
    "city": "城市",
    "management_account_name": "管理科目名称",
    "category": "科目类别",
    "financial_account_code": "财务科目编码",
}


def _source_header(field_id: str, display_name: str) -> str:
    return FIELD_ALIASES.get(field_id, f"{display_name}(源)")


def _external_value(field_id: str, value: Any, row_index: int) -> Any:
    if value is None:
        return None
    if field_id.endswith("month") or field_id == "month_key":
        text = str(value)
        year, month = text[:7].split("-")
        return f"{year}/{month}" if row_index % 2 == 0 else f"{year}年{month}月"
    if isinstance(value, Decimal):
        text = format(value, "f")
        if row_index % 7 == 0:
            whole, separator, fraction = text.partition(".")
            grouped = f"{int(whole):,}"
            return grouped + (separator + fraction if separator else "")
        return text
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def _ordered_fields(fields: tuple[Any, ...]) -> tuple[Any, ...]:
    if len(fields) < 3:
        return fields
    return fields[2:] + fields[:2]


def _display_width(value: Any) -> int:
    text = "" if value is None else str(value)
    return sum(
        2 if unicodedata.east_asian_width(character) in {"F", "W"} else 1
        for character in text
    )


def _normalize_xlsx(path: Path) -> None:
    descriptor, temporary_name = tempfile.mkstemp(suffix=".xlsx", dir=path.parent)
    os.close(descriptor)
    temporary = Path(temporary_name)
    try:
        with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(
            temporary, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9
        ) as destination:
            for name in sorted(source.namelist()):
                info = zipfile.ZipInfo(name, date_time=ZIP_TIME)
                info.compress_type = zipfile.ZIP_DEFLATED
                info.create_system = 3
                info.external_attr = 0o600 << 16
                content = source.read(name)
                if name == "docProps/core.xml":
                    content = re.sub(
                        rb"<dcterms:modified[^>]*>[^<]*</dcterms:modified>",
                        (
                            b'<dcterms:modified xsi:type="dcterms:W3CDTF">'
                            b"2026-08-30T00:00:00Z</dcterms:modified>"
                        ),
                        content,
                    )
                destination.writestr(info, content)
        os.replace(temporary, path)
        path.chmod(0o644)
    finally:
        temporary.unlink(missing_ok=True)


def generate() -> dict[str, Any]:
    contract = load_contract(CONTRACT_PATH)
    package = build_reference_package()
    rows_by_role = workbook_rows(package)
    sheets_by_role = {sheet.sheet_id: sheet for sheet in contract.sheets}

    workbook = Workbook()
    active_sheet = workbook.active
    if active_sheet is None:
        raise RuntimeError("new workbook did not contain an active worksheet")
    workbook.remove(active_sheet)
    workbook.properties.creator = "FLOW"
    workbook.properties.title = "外部物流经营数据样例"
    workbook.properties.description = "用于验证非标准 Excel 识别、映射和清洗"
    workbook.properties.created = FIXED_TIME
    workbook.properties.modified = FIXED_TIME

    manifest_sheets: dict[str, Any] = {}
    roles: dict[str, str] = {}

    for sheet_name, role in SHEET_ORDER:
        sheet_contract = sheets_by_role[role]
        source_rows = rows_by_role[role]
        ordered_fields = _ordered_fields(sheet_contract.fields)
        aliases = {
            field.field_id: _source_header(field.field_id, field.display_name)
            for field in ordered_fields
        }
        worksheet = workbook.create_sheet(sheet_name)
        last_column = get_column_letter(len(ordered_fields) + 1)
        worksheet.merge_cells(f"A1:{last_column}1")
        worksheet["A1"] = "FLOW 外部物流经营数据样例"
        worksheet["A1"].font = Font(bold=True, color="FFFFFF", size=14)
        worksheet["A1"].fill = PatternFill("solid", fgColor="1F4E78")
        worksheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
        worksheet["A2"] = f"来源系统导出：{sheet_name}；第 3 行为表头，附加备注列无需导入。"
        worksheet.merge_cells(f"A2:{last_column}2")
        worksheet["A2"].font = Font(color="666666", italic=True)

        headers = [aliases[field.field_id] for field in ordered_fields] + ["数据备注"]
        worksheet.append(headers)
        for cell in worksheet[3]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="5B9BD5")
            cell.alignment = Alignment(wrap_text=True, vertical="center")

        for row_offset, source_row in enumerate(source_rows, start=1):
            values = [
                _external_value(field.field_id, source_row.get(field.field_id), row_offset)
                for field in ordered_fields
            ]
            values.append("系统导出" if row_offset % 97 == 0 else None)
            worksheet.append(values)

        worksheet.freeze_panes = "A4"
        worksheet.auto_filter.ref = f"A3:{last_column}{worksheet.max_row}"
        worksheet.sheet_view.showGridLines = False
        worksheet.row_dimensions[1].height = 28
        worksheet.row_dimensions[3].height = 28
        for column_index, header in enumerate(headers, start=1):
            sample_width = max(
                _display_width(worksheet.cell(row=row, column=column_index).value)
                for row in range(3, min(worksheet.max_row, 120) + 1)
            )
            maximum_width = 40 if header == "来源记录号" else 30
            worksheet.column_dimensions[get_column_letter(column_index)].width = min(
                max(sample_width + 3, 14), maximum_width
            )

        roles[role] = sheet_name
        manifest_sheets[sheet_name] = {
            "expected_role": role,
            "header_row": 3,
            "data_start_row": 4,
            "row_count": len(source_rows),
            "field_aliases": aliases,
            "ignored_columns": ["数据备注"],
        }

    WORKBOOK_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(WORKBOOK_PATH)
    _normalize_xlsx(WORKBOOK_PATH)

    manifest = {
        "fixture_id": "external.logistics.nonstandard.v1",
        "expected_contract_version": contract.contract_version,
        "workbook_sha256": hashlib.sha256(WORKBOOK_PATH.read_bytes()).hexdigest(),
        "sheet_order": [name for name, _ in SHEET_ORDER],
        "roles": roles,
        "sheets": manifest_sheets,
        "intentional_variations": [
            "renamed_and_reordered_sheets",
            "two_preamble_rows",
            "aliased_and_reordered_columns",
            "mixed_month_formats",
            "comma_formatted_decimal_strings",
            "ignorable_note_column",
            "no_flow_field_id_row",
        ],
    }
    MANIFEST_PATH.parent.mkdir(parents=True, exist_ok=True)
    MANIFEST_PATH.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    return manifest


if __name__ == "__main__":
    generate()
