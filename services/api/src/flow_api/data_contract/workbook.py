from __future__ import annotations

import hashlib
import json
from collections.abc import Mapping
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from flow_api.data_contract.models import FieldContract, SheetContract, WorkbookContract
from flow_api.data_contract.records import CanonicalPackage

FLOW_NAVY = "172A46"
FLOW_BLUE = "2F6BFF"
FLOW_LIGHT_BLUE = "EAF1FF"
FLOW_PALE = "F5F7FA"
FLOW_REQUIRED = "FFF2CC"
FIXED_WORKBOOK_TIME = datetime(2026, 8, 30, 0, 0, 0)


def _instructions() -> tuple[dict[str, str | None], ...]:
    return (
        {
            "section_code": "PURPOSE",
            "section_name": "模板用途",
            "instruction": (
                "本工作簿既可由业务和财务人员直接填写，也可作为外部数据转换后的统一核对格式。"
            ),
            "example": "先选择分析场景，再填写或转换对应工作表。",
        },
        {
            "section_code": "NON_STANDARD",
            "section_name": "非标准数据",
            "instruction": (
                "来自 ERP、业务系统或非标准 Excel 的数据必须先映射到本契约，"
                "再进入指标、驾驶舱和报告。"
            ),
            "example": "客户名称列可映射到 customer_name，但客户编码必须确认。",
        },
        {
            "section_code": "FIELD_ID",
            "section_name": "字段 ID",
            "instruction": (
                "第 2 行是不可修改的稳定字段 ID；系统依赖字段 ID 而不是列位置或第 1 行显示名称。"
            ),
            "example": "revenue 始终代表营业收入，即使移动列或修改显示名称。",
        },
        {
            "section_code": "DATA_START",
            "section_name": "数据区域",
            "instruction": (
                "第 1 行为显示名称，第 2 行为字段 ID，第 3 行为类型和单位提示，数据从第 4 行开始。"
            ),
            "example": "不要在数据区域插入第二套表头或小计行。",
        },
        {
            "section_code": "BLOCKING",
            "section_name": "阻断错误",
            "instruction": (
                "必填缺失、稳定主键重复、关系断裂、契约版本不兼容和经营财务对账超差将阻止发布。"
            ),
            "example": "customer_code 在客户主数据中不存在。",
        },
        {
            "section_code": "WARNING",
            "section_name": "确认警告",
            "instruction": (
                "异常负值、零订单有收入、单位疑似错误等需要 Finance BP 确认，但不必然阻断修订。"
            ),
            "example": "订单量为 0 但营业收入不为 0。",
        },
        {
            "section_code": "RECONCILIATION",
            "section_name": "跨表对账",
            "instruction": (
                "经营实际中的收入和直接成本必须与财务实际中的对应管理科目在配置阈值内一致。"
            ),
            "example": "02_经营实际收入合计 = 03_财务实际 REVENUE 合计。",
        },
    )


def workbook_rows(package: CanonicalPackage) -> dict[str, tuple[Mapping[str, Any], ...]]:
    segment_names = {segment.code: segment.name for segment in package.customer_segments}
    customer_rows = tuple(
        {
            "customer_code": customer.code,
            "customer_name": customer.name,
            "industry": customer.industry,
            "tier": customer.tier,
            "credit_term_days": customer.credit_term_days,
            "customer_segment_code": customer.segment_code,
            "customer_segment_name": segment_names[customer.segment_code],
        }
        for customer in package.customers
    )
    product_rows = tuple(
        {
            "logistics_product_code": product.code,
            "logistics_product_name": product.name,
            "level": product.level,
            "parent_code": product.parent_code,
        }
        for product in package.logistics_products
    )
    organization_rows = tuple(
        {
            "entity_type": "organization",
            "entity_code": organization.code,
            "entity_name": organization.name,
            "level": organization.level,
            "parent_code": organization.parent_code,
            "province": None,
            "city": None,
        }
        for organization in package.organizations
    )
    region_rows = tuple(
        {
            "entity_type": "region",
            "entity_code": region.code,
            "entity_name": region.name,
            "level": "region",
            "parent_code": region.parent_code,
            "province": region.province,
            "city": region.city,
        }
        for region in package.regions
    )
    account_rows = tuple(
        {
            "management_account_code": account.code,
            "management_account_name": account.name,
            "category": account.category,
            "financial_account_code": account.financial_account_code,
            "parent_code": account.parent_code,
        }
        for account in package.management_accounts
    )
    return {
        "instructions": _instructions(),
        "analysis_batch": (package.batch.model_dump(),),
        "operating_actual": tuple(row.model_dump() for row in package.operating_actuals),
        "financial_actual": tuple(row.model_dump() for row in package.financial_actuals),
        "monthly_budget": tuple(row.model_dump() for row in package.monthly_budgets),
        "ar_collection": tuple(row.model_dump() for row in package.ar_collections),
        "customer_master": customer_rows,
        "logistics_product": product_rows,
        "organization_region": organization_rows + region_rows,
        "management_account": account_rows,
    }


def _cell_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def _hint(field: FieldContract) -> str:
    requirement = "必填" if field.required else "可选"
    unit = f" / {field.unit}" if field.unit else ""
    nullable = " / 可空" if field.nullable else ""
    return f"{field.data_type}{unit} / {requirement}{nullable}"


def _number_format(field: FieldContract) -> str:
    if field.data_type == "decimal":
        return "#,##0." + "0" * (field.scale or 0)
    if field.data_type == "integer":
        return "0"
    return "General"


def _add_validation(
    worksheet: Any, field: FieldContract, column_letter: str, maximum_row: int
) -> None:
    cell_range = f"{column_letter}4:{column_letter}{maximum_row}"
    validation: DataValidation | None = None
    if field.data_type == "enum" and field.enum:
        validation = DataValidation(
            type="list",
            formula1='"' + ",".join(field.enum) + '"',
            allow_blank=field.nullable,
        )
    elif field.data_type == "decimal":
        validation = DataValidation(
            type="decimal",
            operator="greaterThanOrEqual" if field.minimum is not None else "between",
            formula1=field.minimum or "-999999999999999999",
            formula2=None if field.minimum is not None else "999999999999999999",
            allow_blank=field.nullable,
        )
    elif field.data_type == "integer":
        validation = DataValidation(
            type="whole",
            operator="greaterThanOrEqual" if field.minimum is not None else "between",
            formula1=field.minimum or "-999999999",
            formula2=None if field.minimum is not None else "999999999",
            allow_blank=field.nullable,
        )
    elif field.data_type == "month":
        validation = DataValidation(
            type="textLength", operator="equal", formula1="7", allow_blank=field.nullable
        )
    if validation is not None:
        validation.errorTitle = "FLOW 数据格式不符合契约"
        validation.error = field.description
        validation.promptTitle = field.display_name
        validation.prompt = _hint(field)
        validation.showErrorMessage = True
        validation.showInputMessage = True
        worksheet.add_data_validation(validation)
        validation.add(cell_range)


def _style_sheet(worksheet: Any, sheet: SheetContract, row_count: int) -> None:
    last_column = get_column_letter(len(sheet.fields))
    last_row = max(row_count + 3, 4)
    worksheet.freeze_panes = "A4"
    worksheet.auto_filter.ref = f"A1:{last_column}{last_row}"
    worksheet.row_dimensions[1].height = 30
    worksheet.row_dimensions[2].height = 24
    worksheet.row_dimensions[3].height = 22

    for column_index, field in enumerate(sheet.fields, start=1):
        display_cell = worksheet.cell(row=1, column=column_index)
        field_id_cell = worksheet.cell(row=2, column=column_index)
        hint_cell = worksheet.cell(row=3, column=column_index)
        display_cell.fill = PatternFill("solid", fgColor=FLOW_NAVY)
        display_cell.font = Font(color="FFFFFF", bold=True, size=11)
        display_cell.alignment = Alignment(vertical="center", wrap_text=True)
        display_cell.comment = Comment(field.description, "FLOW")
        field_id_cell.fill = PatternFill("solid", fgColor=FLOW_LIGHT_BLUE)
        field_id_cell.font = Font(color=FLOW_BLUE, bold=True, size=9)
        field_id_cell.protection = Protection(locked=True)
        hint_cell.fill = PatternFill(
            "solid", fgColor=FLOW_REQUIRED if field.required else FLOW_PALE
        )
        hint_cell.font = Font(color="6B7280", italic=True, size=9)
        hint_cell.alignment = Alignment(wrap_text=True)
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(len(field.display_name) * 2 + 4, len(field.field_id) + 3, 14), 34
        )
        for row_index in range(4, last_row + 1):
            data_cell = worksheet.cell(row=row_index, column=column_index)
            data_cell.protection = Protection(locked=False)
            data_cell.number_format = _number_format(field)
        _add_validation(
            worksheet,
            field,
            get_column_letter(column_index),
            last_row + 1000,
        )

    worksheet.protection.sheet = True
    worksheet.protection.autoFilter = False
    worksheet.protection.sort = False
    worksheet.protection.insertRows = False
    worksheet.sheet_view.showGridLines = False
    if sheet.sheet_id == "instructions":
        instruction_widths = {"A": 22, "B": 24, "C": 78, "D": 60}
        for column_letter, width in instruction_widths.items():
            worksheet.column_dimensions[column_letter].width = width
        for row_index in range(4, last_row + 1):
            worksheet.row_dimensions[row_index].height = 46
            for cell in worksheet[row_index]:
                cell.alignment = Alignment(vertical="top", wrap_text=True)


def render_workbook(
    contract: WorkbookContract, package: CanonicalPackage, destination: Path
) -> None:
    if package.batch.contract_version != contract.contract_version:
        raise ValueError(
            f"package contract {package.batch.contract_version} does not match "
            f"{contract.contract_version}"
        )
    rows_by_sheet = workbook_rows(package)
    workbook = Workbook()
    active_sheet = workbook.active
    if active_sheet is None:
        raise RuntimeError("new workbook did not create an active worksheet")
    workbook.remove(active_sheet)
    workbook.properties.creator = "FLOW"
    workbook.properties.title = contract.workbook_name
    workbook.properties.subject = "Finance BP 经营分析标准数据契约"
    workbook.properties.description = "FLOW 标准 Excel 是统一数据中间层的外部交换格式。"
    workbook.properties.keywords = "FLOW,Finance BP,flow.excel.v1"
    workbook.properties.version = contract.contract_version
    workbook.properties.created = FIXED_WORKBOOK_TIME
    workbook.properties.modified = FIXED_WORKBOOK_TIME

    for sheet in contract.sheets:
        worksheet = workbook.create_sheet(sheet.sheet_name)
        rows = rows_by_sheet[sheet.sheet_id]
        worksheet.append([field.display_name for field in sheet.fields])
        worksheet.append([field.field_id for field in sheet.fields])
        worksheet.append([_hint(field) for field in sheet.fields])
        for row in rows:
            worksheet.append([_cell_value(row.get(field.field_id)) for field in sheet.fields])
        _style_sheet(worksheet, sheet, len(rows))

    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)


def workbook_semantic_fingerprint(path: Path) -> str:
    workbook = load_workbook(path, data_only=False, read_only=False)
    payload: list[dict[str, Any]] = []
    for worksheet in workbook.worksheets:
        rows = [
            [
                cell.value.isoformat() if isinstance(cell.value, datetime) else cell.value
                for cell in row
            ]
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row)
        ]
        validations = sorted(
            (
                validation.type,
                validation.operator,
                validation.formula1,
                validation.formula2,
                str(validation.sqref),
            )
            for validation in worksheet.data_validations.dataValidation
        )
        payload.append(
            {
                "title": worksheet.title,
                "rows": rows,
                "freeze_panes": str(worksheet.freeze_panes),
                "auto_filter": worksheet.auto_filter.ref,
                "validations": validations,
            }
        )
    content = json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str).encode()
    return hashlib.sha256(content).hexdigest()
