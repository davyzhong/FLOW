from __future__ import annotations

import re
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict

from flow_api.data_contract.models import FieldContract, SheetContract, WorkbookContract
from flow_api.data_contract.records import (
    ArCollectionRecord,
    BatchRecord,
    CanonicalPackage,
    CustomerRecord,
    CustomerSegmentRecord,
    FinancialActualRecord,
    LogisticsProductRecord,
    ManagementAccountRecord,
    MonthlyBudgetRecord,
    OperatingActualRecord,
    OrganizationRecord,
    PeriodRecord,
    RegionRecord,
    ScenarioVersionRecord,
)

MONTH_PATTERN = re.compile(r"^[0-9]{4}-(0[1-9]|1[0-2])$")


class ParseIssue(BaseModel):
    model_config = ConfigDict(extra="forbid", frozen=True)

    sheet_name: str
    row: int
    field_id: str | None = None
    code: str
    severity: str = "blocking"
    message: str


class WorkbookParseError(ValueError):
    def __init__(self, issues: tuple[ParseIssue, ...]) -> None:
        self.issues = issues
        super().__init__(f"workbook contains {len(issues)} blocking issue(s)")


def _invalid_type(field: FieldContract, value: Any) -> ValueError:
    return ValueError(
        f"{field.field_id} expects {field.data_type}, received {type(value).__name__}"
    )


def _convert_value(field: FieldContract, value: Any) -> Any:
    if value is None:
        if field.nullable:
            return None
        raise ValueError(f"{field.field_id} is required and cannot be blank")
    if field.data_type in {"string", "enum"}:
        if not isinstance(value, str):
            raise _invalid_type(field, value)
        if field.data_type == "enum" and field.enum is not None and value not in field.enum:
            raise ValueError(f"{field.field_id} is not an allowed enum value: {value}")
        return value
    if field.data_type == "month":
        if not isinstance(value, str) or MONTH_PATTERN.fullmatch(value) is None:
            raise _invalid_type(field, value)
        return value
    if field.data_type == "datetime":
        if isinstance(value, datetime):
            return value
        if not isinstance(value, str):
            raise _invalid_type(field, value)
        try:
            return datetime.fromisoformat(value.replace("Z", "+00:00"))
        except ValueError as error:
            raise _invalid_type(field, value) from error
    if field.data_type == "integer":
        if isinstance(value, bool):
            raise _invalid_type(field, value)
        if isinstance(value, int):
            result = value
        elif isinstance(value, float) and value.is_integer():
            result = int(value)
        else:
            raise _invalid_type(field, value)
        if field.minimum is not None and result < int(field.minimum):
            raise ValueError(f"{field.field_id} is below minimum {field.minimum}")
        return result
    if field.data_type == "decimal":
        if isinstance(value, bool):
            raise _invalid_type(field, value)
        try:
            decimal_value = value if isinstance(value, Decimal) else Decimal(str(value))
        except (InvalidOperation, ValueError) as error:
            raise _invalid_type(field, value) from error
        quantum = Decimal(1).scaleb(-(field.scale or 0))
        quantized = decimal_value.quantize(quantum)
        if quantized != decimal_value:
            raise ValueError(
                f"{field.field_id} exceeds declared scale {field.scale}: {decimal_value}"
            )
        if field.minimum is not None and quantized < Decimal(field.minimum):
            raise ValueError(f"{field.field_id} is below minimum {field.minimum}")
        return quantized
    raise _invalid_type(field, value)


def _parse_sheet(
    worksheet: Any, sheet: SheetContract, issues: list[ParseIssue]
) -> tuple[dict[str, Any], ...]:
    raw_field_ids = [
        worksheet.cell(row=2, column=index).value for index in range(1, worksheet.max_column + 1)
    ]
    known_ids = {field.field_id for field in sheet.fields}
    present_ids = [value for value in raw_field_ids if isinstance(value, str)]
    duplicate_ids = {field_id for field_id in present_ids if present_ids.count(field_id) > 1}
    for field_id in sorted(duplicate_ids):
        issues.append(
            ParseIssue(
                sheet_name=sheet.sheet_name,
                row=2,
                field_id=field_id,
                code="duplicate_field_id",
                message=f"field ID appears more than once: {field_id}",
            )
        )
    for field_id in sorted(set(present_ids) - known_ids):
        issues.append(
            ParseIssue(
                sheet_name=sheet.sheet_name,
                row=2,
                field_id=field_id,
                code="unknown_field_id",
                message=f"field ID is not in {sheet.sheet_id}: {field_id}",
            )
        )
    for field in sheet.fields:
        if field.field_id not in present_ids:
            issues.append(
                ParseIssue(
                    sheet_name=sheet.sheet_name,
                    row=2,
                    field_id=field.field_id,
                    code="missing_field_id",
                    message=f"required contract column is absent: {field.field_id}",
                )
            )
    if duplicate_ids or set(present_ids) - known_ids or known_ids - set(present_ids):
        return ()

    column_by_field = {field_id: raw_field_ids.index(field_id) + 1 for field_id in present_ids}
    rows: list[dict[str, Any]] = []
    for row_index in range(4, worksheet.max_row + 1):
        raw_values = [
            worksheet.cell(row=row_index, column=column_by_field[field.field_id]).value
            for field in sheet.fields
        ]
        if all(value is None for value in raw_values):
            continue
        parsed: dict[str, Any] = {}
        row_valid = True
        for field, raw_value in zip(sheet.fields, raw_values, strict=True):
            try:
                parsed[field.field_id] = _convert_value(field, raw_value)
            except (ValueError, InvalidOperation) as error:
                row_valid = False
                issues.append(
                    ParseIssue(
                        sheet_name=sheet.sheet_name,
                        row=row_index,
                        field_id=field.field_id,
                        code="invalid_type",
                        message=str(error),
                    )
                )
        if row_valid:
            rows.append(parsed)
    return tuple(rows)


def _validate_grains(
    contract: WorkbookContract,
    rows_by_sheet: dict[str, tuple[dict[str, Any], ...]],
    issues: list[ParseIssue],
) -> None:
    for sheet in contract.sheets:
        if not sheet.grain:
            continue
        seen: dict[tuple[Any, ...], int] = {}
        for offset, row in enumerate(rows_by_sheet[sheet.sheet_id], start=4):
            grain = tuple(row[field_id] for field_id in sheet.grain)
            if grain in seen:
                issues.append(
                    ParseIssue(
                        sheet_name=sheet.sheet_name,
                        row=offset,
                        code="duplicate_grain",
                        message=f"grain duplicates row {seen[grain]}: {grain}",
                    )
                )
            else:
                seen[grain] = offset


def _relationship_issue(
    issues: list[ParseIssue], sheet_name: str, row: int, field_id: str, value: str
) -> None:
    issues.append(
        ParseIssue(
            sheet_name=sheet_name,
            row=row,
            field_id=field_id,
            code="broken_relationship",
            message=f"referenced business key does not exist: {value}",
        )
    )


def _validate_relationships(
    rows: dict[str, tuple[dict[str, Any], ...]], issues: list[ParseIssue]
) -> None:
    organization_codes = {
        row["entity_code"]
        for row in rows["organization_region"]
        if row["entity_type"] == "organization"
    }
    region_codes = {
        row["entity_code"] for row in rows["organization_region"] if row["entity_type"] == "region"
    }
    customer_codes = {row["customer_code"] for row in rows["customer_master"]}
    segment_codes = {row["customer_segment_code"] for row in rows["customer_master"]}
    product_codes = {row["logistics_product_code"] for row in rows["logistics_product"]}
    account_codes = {row["management_account_code"] for row in rows["management_account"]}
    checks: tuple[tuple[str, str, set[Any]], ...] = (
        ("operating_actual", "organization_code", organization_codes),
        ("operating_actual", "customer_code", customer_codes),
        ("operating_actual", "logistics_product_code", product_codes),
        ("operating_actual", "region_code", region_codes),
        ("financial_actual", "organization_code", organization_codes),
        ("financial_actual", "management_account_code", account_codes),
        ("monthly_budget", "organization_code", organization_codes),
        ("monthly_budget", "customer_segment_code", segment_codes),
        ("monthly_budget", "logistics_product_code", product_codes),
        ("monthly_budget", "management_account_code", account_codes),
        ("ar_collection", "customer_code", customer_codes),
    )
    sheet_names = {
        "operating_actual": "02_经营实际",
        "financial_actual": "03_财务实际",
        "monthly_budget": "04_月度预算",
        "ar_collection": "05_应收回款",
    }
    for sheet_id, field_id, targets in checks:
        for offset, row in enumerate(rows[sheet_id], start=4):
            value = row[field_id]
            if value is not None and value not in targets:
                _relationship_issue(issues, sheet_names[sheet_id], offset, field_id, str(value))


def _build_package(rows: dict[str, tuple[dict[str, Any], ...]]) -> CanonicalPackage:
    batch = BatchRecord.model_validate(rows["analysis_batch"][0])
    customer_segments_by_code: dict[str, CustomerSegmentRecord] = {}
    customers: list[CustomerRecord] = []
    for row in rows["customer_master"]:
        segment_code = str(row["customer_segment_code"])
        customer_segments_by_code[segment_code] = CustomerSegmentRecord(
            code=segment_code, name=str(row["customer_segment_name"])
        )
        customers.append(
            CustomerRecord(
                code=str(row["customer_code"]),
                name=str(row["customer_name"]),
                industry=row["industry"],
                tier=row["tier"],
                credit_term_days=row["credit_term_days"],
                segment_code=segment_code,
            )
        )
    organizations = tuple(
        OrganizationRecord(
            code=str(row["entity_code"]),
            name=str(row["entity_name"]),
            level=str(row["level"]),
            parent_code=row["parent_code"],
        )
        for row in rows["organization_region"]
        if row["entity_type"] == "organization"
    )
    regions = tuple(
        RegionRecord(
            code=str(row["entity_code"]),
            name=str(row["entity_name"]),
            province=row["province"],
            city=row["city"],
            parent_code=row["parent_code"],
        )
        for row in rows["organization_region"]
        if row["entity_type"] == "region"
    )
    products = tuple(
        LogisticsProductRecord(
            code=str(row["logistics_product_code"]),
            name=str(row["logistics_product_name"]),
            level=str(row["level"]),
            parent_code=row["parent_code"],
        )
        for row in rows["logistics_product"]
    )
    accounts = tuple(
        ManagementAccountRecord(
            code=str(row["management_account_code"]),
            name=str(row["management_account_name"]),
            category=str(row["category"]),
            financial_account_code=row["financial_account_code"],
            parent_code=row["parent_code"],
        )
        for row in rows["management_account"]
    )
    operating = tuple(OperatingActualRecord.model_validate(row) for row in rows["operating_actual"])
    financial = tuple(FinancialActualRecord.model_validate(row) for row in rows["financial_actual"])
    budgets = tuple(MonthlyBudgetRecord.model_validate(row) for row in rows["monthly_budget"])
    ar_rows = tuple(ArCollectionRecord.model_validate(row) for row in rows["ar_collection"])
    all_months = sorted(
        {row.month_key for row in operating}
        | {row.month_key for row in financial}
        | {row.month_key for row in budgets}
        | {row.month_key for row in ar_rows}
    )
    periods = tuple(
        PeriodRecord(
            month_key=month_key,
            year=int(month_key[:4]),
            month=int(month_key[5:]),
            quarter=(int(month_key[5:]) - 1) // 3 + 1,
            window=(
                "comparison"
                if batch.comparison_start_month <= month_key <= batch.comparison_end_month
                else "analysis"
            ),
        )
        for month_key in all_months
    )
    return CanonicalPackage(
        batch=batch,
        periods=periods,
        organizations=organizations,
        customer_segments=tuple(
            customer_segments_by_code[code] for code in sorted(customer_segments_by_code)
        ),
        customers=tuple(customers),
        logistics_products=products,
        regions=regions,
        management_accounts=accounts,
        scenario_versions=(
            ScenarioVersionRecord(
                code=batch.actual_scenario_code,
                name="实际",
                scenario_type="actual",
            ),
            ScenarioVersionRecord(
                code=batch.budget_scenario_code,
                name="FY26 月度预算",
                scenario_type="budget",
                version_label=batch.budget_version_label,
            ),
        ),
        operating_actuals=operating,
        financial_actuals=financial,
        monthly_budgets=budgets,
        ar_collections=ar_rows,
    )


def parse_workbook(path: Path, contract: WorkbookContract) -> CanonicalPackage:
    workbook = load_workbook(path, data_only=False, read_only=False)
    issues: list[ParseIssue] = []
    expected_names = {sheet.sheet_name for sheet in contract.sheets}
    actual_names = set(workbook.sheetnames)
    for missing_name in sorted(expected_names - actual_names):
        issues.append(
            ParseIssue(
                sheet_name=missing_name,
                row=0,
                code="missing_sheet",
                message=f"required sheet is absent: {missing_name}",
            )
        )
    if issues:
        raise WorkbookParseError(tuple(issues))

    rows_by_sheet = {
        sheet.sheet_id: _parse_sheet(workbook[sheet.sheet_name], sheet, issues)
        for sheet in contract.sheets
    }
    _validate_grains(contract, rows_by_sheet, issues)
    _validate_relationships(rows_by_sheet, issues)
    batch_rows = rows_by_sheet["analysis_batch"]
    if len(batch_rows) != 1:
        issues.append(
            ParseIssue(
                sheet_name="01_分析批次",
                row=4,
                code="invalid_batch_count",
                message="analysis batch sheet must contain exactly one data row",
            )
        )
    elif batch_rows[0]["contract_version"] != contract.contract_version:
        issues.append(
            ParseIssue(
                sheet_name="01_分析批次",
                row=4,
                field_id="contract_version",
                code="incompatible_contract_version",
                message=(
                    f"expected {contract.contract_version}, received "
                    f"{batch_rows[0]['contract_version']}"
                ),
            )
        )
    if issues:
        raise WorkbookParseError(tuple(issues))
    return _build_package(rows_by_sheet)
