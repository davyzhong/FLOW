"""空白 flow.excel.v1 模板渲染器（确定性字节输出）。

模板与 render_workbook 共享冻结契约：说明页 + 每张工作表的前三行契约行
（显示名 / 字段 ID / 填写提示），数据区留空供 Finance BP 填写。
所有属性使用 FIXED_WORKBOOK_TIME，保证同契约两次渲染字节稳定。
"""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from flow_api.data_contract.workbook import FIXED_WORKBOOK_TIME, _hint

TEMPLATE_ID = "flow.excel.v1"
TEMPLATE_FILENAME = "flow.excel.v1.template.xlsx"
TEMPLATE_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_INSTRUCTIONS = (
    ("FLOW 标准 Excel 数据包模板（flow.excel.v1）", True),
    ("", False),
    ("填写说明：", True),
    ("1. 请勿修改前两行的字段显示名与字段 ID，也不要删除第三行的填写提示。", False),
    ("2. 请从第 4 行开始填写数据；不要插入额外的表头或合并单元格。", False),
    ("3. 日期统一为 YYYY-MM-DD；金额单位为元，保留两位小数。", False),
    ("4. 完整字段口径见模板同目录的契约文件与 FLOW 使用手册。", False),
)


def render_blank_template(contract: Any) -> bytes:
    workbook = Workbook()
    active = workbook.active
    if active is None:
        raise RuntimeError("new workbook did not create an active worksheet")
    workbook.remove(active)
    workbook.properties.creator = "FLOW"
    workbook.properties.title = f"{contract.workbook_name} blank template"
    workbook.properties.subject = "Finance BP 经营分析标准数据契约"
    workbook.properties.description = "FLOW 标准 Excel 空白模板，供 Finance BP 下载填写。"
    workbook.properties.keywords = "FLOW,Finance BP,flow.excel.v1,template"
    workbook.properties.version = contract.contract_version
    workbook.properties.created = FIXED_WORKBOOK_TIME
    workbook.properties.modified = FIXED_WORKBOOK_TIME

    instructions = workbook.create_sheet("说明")
    for row, (text, bold) in enumerate(_INSTRUCTIONS, start=1):
        cell = instructions.cell(row=row, column=1, value=text)
        cell.font = Font(bold=bold)
        cell.alignment = Alignment(wrap_text=False)
    instructions.column_dimensions["A"].width = 96
    instructions.sheet_view.showGridLines = False

    header_fill = PatternFill("solid", fgColor="E8F0FE")
    for sheet in contract.sheets:
        worksheet = workbook.create_sheet(sheet.sheet_name)
        worksheet.cell(row=1, column=1, value="字段显示名")
        worksheet.cell(row=2, column=1, value="字段 ID")
        worksheet.cell(row=3, column=1, value="填写提示")
        for index, field in enumerate(sheet.fields, start=1):
            column = get_column_letter(index)
            display = worksheet.cell(row=1, column=index, value=field.display_name)
            field_id = worksheet.cell(row=2, column=index, value=field.field_id)
            hint = worksheet.cell(row=3, column=index, value=_hint(field))
            for cell in (display, field_id, hint):
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="left")
            worksheet.column_dimensions[column].width = max(14, len(field.display_name) * 2)
        worksheet.freeze_panes = "A4"
    workbook.active = workbook["说明"]

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
