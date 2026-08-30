from __future__ import annotations

import hashlib
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any, Literal

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from flow_api.data_contract.models import WorkbookContract
from flow_api.data_contract.parser import build_package_from_rows
from flow_api.data_contract.records import CanonicalPackage
from flow_api.intake.mapping import MappingProposal
from flow_api.intake.models import WorkbookProfile
from flow_api.intake.transforms import TransformError, TransformRegistry, apply_transform

LineageStatus = Literal["unchanged", "transformed", "failed"]


@dataclass(frozen=True, slots=True)
class LineageValue:
    source_sheet: str
    source_row: int
    source_column: str
    source_header: str
    target_sheet_id: str
    target_field_id: str
    raw_value: Any
    transformed_value: Any
    rule_id: str
    rule_version: int
    status: LineageStatus
    reason: str


@dataclass(frozen=True, slots=True)
class ExtractedCandidate:
    source_sha256: str
    package: CanonicalPackage
    lineage: tuple[LineageValue, ...]

    @property
    def failed_lineage(self) -> tuple[LineageValue, ...]:
        return tuple(item for item in self.lineage if item.status == "failed")


class CandidateExtractionError(ValueError):
    def __init__(self, failed_lineage: tuple[LineageValue, ...]) -> None:
        self.failed_lineage = failed_lineage
        super().__init__(f"candidate extraction has {len(failed_lineage)} failed values")


def _read_source(source: Path | bytes) -> bytes:
    return source.read_bytes() if isinstance(source, Path) else bytes(source)


def extract_candidate_package(
    source: Path | bytes,
    profile: WorkbookProfile,
    proposal: MappingProposal,
    contract: WorkbookContract,
    transforms: TransformRegistry,
) -> ExtractedCandidate:
    content = _read_source(source)
    source_sha256 = hashlib.sha256(content).hexdigest()
    if source_sha256 != profile.sha256 or source_sha256 != proposal.source_sha256:
        raise ValueError("source bytes, workbook profile and mapping proposal do not match")
    if proposal.unresolved_sheet_ids or any(
        sheet.unresolved_required_fields for sheet in proposal.sheets
    ):
        raise ValueError("mapping proposal still has unresolved required identities")

    workbook = load_workbook(BytesIO(content), read_only=True, data_only=False)
    rows_by_sheet: dict[str, tuple[dict[str, Any], ...]] = {}
    lineage: list[LineageValue] = []
    try:
        for sheet_mapping in proposal.sheets:
            source_profile = profile.get_sheet(sheet_mapping.source_sheet)
            if source_profile.data_start_row is None or source_profile.data_end_row is None:
                rows_by_sheet[sheet_mapping.target_sheet_id] = ()
                continue
            worksheet = workbook[sheet_mapping.source_sheet]
            target_sheet = contract.get_sheet(sheet_mapping.target_sheet_id)
            mappings = {field.target_field_id: field for field in sheet_mapping.fields}
            column_indexes = {
                field_id: column_index_from_string(mapping.source_column)
                for field_id, mapping in mappings.items()
            }
            maximum_column = max(column_indexes.values())
            extracted_rows: list[dict[str, Any]] = []
            source_rows = worksheet.iter_rows(
                min_row=source_profile.data_start_row,
                max_row=source_profile.data_end_row,
                min_col=1,
                max_col=maximum_column,
            )
            for row_number, source_row in enumerate(
                source_rows, start=source_profile.data_start_row
            ):
                raw_by_field = {
                    field_id: source_row[column_indexes[field_id] - 1].value
                    for field_id in mappings
                }
                if all(value is None or not str(value).strip() for value in raw_by_field.values()):
                    continue
                candidate_row: dict[str, Any] = {}
                for target_field in target_sheet.fields:
                    mapping = mappings.get(target_field.field_id)
                    if mapping is None:
                        candidate_row[target_field.field_id] = None
                        continue
                    raw_value = raw_by_field[target_field.field_id]
                    rule = transforms.for_field(target_field)
                    try:
                        result = apply_transform(
                            rule,
                            target_field,
                            raw_value,
                            null_rule_id=transforms.null_rule_id,
                            null_rule_version=transforms.null_rule_version,
                        )
                    except TransformError as error:
                        lineage.append(
                            LineageValue(
                                source_sheet=sheet_mapping.source_sheet,
                                source_row=row_number,
                                source_column=mapping.source_column,
                                source_header=mapping.source_header,
                                target_sheet_id=sheet_mapping.target_sheet_id,
                                target_field_id=target_field.field_id,
                                raw_value=raw_value,
                                transformed_value=None,
                                rule_id=rule.rule_id,
                                rule_version=rule.version,
                                status="failed",
                                reason=str(error),
                            )
                        )
                        continue
                    candidate_row[target_field.field_id] = result.value
                    lineage.append(
                        LineageValue(
                            source_sheet=sheet_mapping.source_sheet,
                            source_row=row_number,
                            source_column=mapping.source_column,
                            source_header=mapping.source_header,
                            target_sheet_id=sheet_mapping.target_sheet_id,
                            target_field_id=target_field.field_id,
                            raw_value=raw_value,
                            transformed_value=result.value,
                            rule_id=result.rule_id,
                            rule_version=result.rule_version,
                            status=result.status,
                            reason=result.reason,
                        )
                    )
                extracted_rows.append(candidate_row)
            rows_by_sheet[sheet_mapping.target_sheet_id] = tuple(extracted_rows)
    finally:
        workbook.close()

    failed = tuple(item for item in lineage if item.status == "failed")
    if failed:
        raise CandidateExtractionError(failed)
    return ExtractedCandidate(
        source_sha256=source_sha256,
        package=build_package_from_rows(rows_by_sheet),
        lineage=tuple(lineage),
    )
