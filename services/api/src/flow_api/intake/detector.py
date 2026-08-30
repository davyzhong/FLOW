from __future__ import annotations

import hashlib
import re
import zipfile
from collections import Counter
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell import Cell, ReadOnlyCell
from openpyxl.utils import get_column_letter

from flow_api.intake.models import (
    CellRegion,
    ColumnProfile,
    PrimitiveType,
    SheetProfile,
    WorkbookProfile,
)

FIELD_ID_PATTERN = re.compile(r"^[a-z][a-z0-9]*(?:_[a-z0-9]+)*$")
MONTH_PATTERN = re.compile(r"^\d{4}(?:-|/|年)(?:0[1-9]|1[0-2])月?$")
INTEGER_PATTERN = re.compile(r"^[+-]?\d+$")
DECIMAL_PATTERN = re.compile(r"^[+-]?(?:\d{1,3}(?:,\d{3})+|\d+)(?:\.\d+)?$")
MAX_ARCHIVE_ENTRIES = 5_000
MAX_UNCOMPRESSED_BYTES = 100 * 1024 * 1024
MAX_COMPRESSION_RATIO = 200
MAX_PROFILE_CELLS = 500_000
MAX_SAMPLE_VALUES = 5


class WorkbookDetectionError(ValueError):
    """Base class for safe workbook profiling failures."""


class InvalidWorkbookError(WorkbookDetectionError):
    """The bytes are not a readable OOXML workbook."""


class UnsafeWorkbookError(WorkbookDetectionError):
    """The workbook exceeds safe structural limits or is encrypted."""


class FormulaInputError(WorkbookDetectionError):
    """A source workbook contains formulas instead of immutable values."""


def _read_source(source: Path | bytes) -> bytes:
    return source.read_bytes() if isinstance(source, Path) else bytes(source)


def _validate_archive(content: bytes) -> None:
    if not zipfile.is_zipfile(BytesIO(content)):
        raise InvalidWorkbookError("source is not an XLSX ZIP archive")
    try:
        with zipfile.ZipFile(BytesIO(content)) as archive:
            entries = archive.infolist()
            if len(entries) > MAX_ARCHIVE_ENTRIES:
                raise UnsafeWorkbookError("workbook contains too many archive entries")
            if any(entry.flag_bits & 0x1 for entry in entries):
                raise UnsafeWorkbookError("encrypted workbooks are not supported")
            total_size = sum(entry.file_size for entry in entries)
            if total_size > MAX_UNCOMPRESSED_BYTES:
                raise UnsafeWorkbookError("workbook uncompressed size exceeds the safety limit")
            for entry in entries:
                if entry.file_size < 1_000_000:
                    continue
                ratio = entry.file_size / max(entry.compress_size, 1)
                if ratio > MAX_COMPRESSION_RATIO:
                    raise UnsafeWorkbookError("workbook archive compression ratio is unsafe")
            if "[Content_Types].xml" not in archive.namelist():
                raise InvalidWorkbookError("archive is not an OOXML workbook")
    except zipfile.BadZipFile as error:
        raise InvalidWorkbookError("source is not a readable XLSX archive") from error


def _nonempty(values: tuple[Any, ...]) -> tuple[Any, ...]:
    return tuple(value for value in values if value is not None and str(value).strip())


def _detect_header(rows: tuple[tuple[Any, ...], ...]) -> tuple[int | None, int | None]:
    for index, row in enumerate(rows, start=1):
        values = _nonempty(row)
        if len(values) < 2 or not all(isinstance(value, str) for value in values):
            continue
        field_id_ratio = sum(bool(FIELD_ID_PATTERN.fullmatch(value)) for value in values) / len(
            values
        )
        if field_id_ratio >= 0.8 and index > 1:
            previous = _nonempty(rows[index - 2])
            if len(previous) >= len(values) * 0.8:
                return index - 1, index

    candidates: list[tuple[float, int]] = []
    for index, row in enumerate(rows, start=1):
        values = _nonempty(row)
        if len(values) < 2 or not all(isinstance(value, str) for value in values):
            continue
        uniqueness = len(set(values)) / len(values)
        density = len(values) / max(len(row), 1)
        candidates.append((len(values) + uniqueness + density - index * 0.01, index))
    if not candidates:
        return None, None
    return max(candidates)[1], None


def _classify(value: Any) -> PrimitiveType:
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, datetime):
        return "datetime"
    if isinstance(value, date):
        return "date"
    if isinstance(value, int):
        return "integer"
    if isinstance(value, (Decimal, float)):
        return "decimal"
    text = str(value).strip()
    if MONTH_PATTERN.fullmatch(text):
        return "month"
    if INTEGER_PATTERN.fullmatch(text):
        return "integer"
    if DECIMAL_PATTERN.fullmatch(text):
        return "decimal"
    return "text"


def _inferred_type(values: list[Any]) -> PrimitiveType:
    if not values:
        return "text"
    counts = Counter(_classify(value) for value in values)
    types = set(counts)
    if types <= {"integer", "decimal"}:
        return "decimal" if "decimal" in types else "integer"
    return counts.most_common(1)[0][0]


def _identity(value: Any) -> tuple[str, str]:
    return type(value).__name__, str(value)


def _profile_sheet(worksheet: Any, index: int) -> SheetProfile:
    max_row = worksheet.max_row or 0
    max_column = worksheet.max_column or 0
    if max_row * max_column > MAX_PROFILE_CELLS:
        raise UnsafeWorkbookError(f"sheet {worksheet.title!r} exceeds the profile cell limit")

    preview = tuple(
        tuple(cell.value for cell in row)
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=min(max_row, 12),
            min_col=1,
            max_col=max_column,
        )
    )
    if not any(_nonempty(row) for row in preview):
        return SheetProfile(
            name=worksheet.title,
            index=index,
            state=worksheet.sheet_state,
            used_region=None,
            header_row=None,
            field_id_row=None,
            data_start_row=None,
            data_end_row=None,
            data_row_count=0,
            columns=(),
            grain_candidates=(),
        )

    header_row, field_id_row = _detect_header(preview)
    if header_row is None:
        return SheetProfile(
            name=worksheet.title,
            index=index,
            state=worksheet.sheet_state,
            used_region=CellRegion(1, max_row, 1, max_column),
            header_row=None,
            field_id_row=None,
            data_start_row=None,
            data_end_row=None,
            data_row_count=0,
            columns=(),
            grain_candidates=(),
        )

    data_start = field_id_row + 2 if field_id_row is not None else header_row + 1
    data_end: int | None = None
    columns_by_index: dict[int, list[Any]] = {}
    for row in worksheet.iter_rows(
        min_row=data_start,
        max_row=max_row,
        min_col=1,
        max_col=max_column,
    ):
        if any(cell.data_type == "f" for cell in row):
            formula_cell = next(cell for cell in row if cell.data_type == "f")
            raise FormulaInputError(
                f"formula input is not allowed at {worksheet.title}!{formula_cell.coordinate}"
            )
        if any(cell.value is not None and str(cell.value).strip() for cell in row):
            data_end = row[0].row
        for column_index, cell in enumerate(row, start=1):
            columns_by_index.setdefault(column_index, []).append(cell.value)

    data_row_count = 0 if data_end is None else data_end - data_start + 1
    header_cells = next(worksheet.iter_rows(min_row=header_row, max_row=header_row))
    field_cells: tuple[Cell | ReadOnlyCell, ...] | None = None
    if field_id_row is not None:
        field_cells = next(worksheet.iter_rows(min_row=field_id_row, max_row=field_id_row))

    columns: list[ColumnProfile] = []
    for cell in header_cells:
        if cell.value is None or not str(cell.value).strip():
            continue
        values = columns_by_index.get(cell.column, [])[:data_row_count]
        non_null = [value for value in values if value is not None and str(value).strip()]
        stable_field_id = None
        if field_cells is not None:
            candidate = field_cells[cell.column - 1].value
            if isinstance(candidate, str) and FIELD_ID_PATTERN.fullmatch(candidate):
                stable_field_id = candidate
        uniqueness_ratio = (
            len({_identity(value) for value in non_null}) / len(non_null) if non_null else 0.0
        )
        columns.append(
            ColumnProfile(
                column_index=cell.column,
                column_letter=get_column_letter(cell.column),
                header=str(cell.value).strip(),
                stable_field_id=stable_field_id,
                inferred_type=_inferred_type(non_null),
                nullable=len(non_null) < data_row_count,
                non_null_count=len(non_null),
                uniqueness_ratio=round(uniqueness_ratio, 6),
                sample_values=tuple(non_null[:MAX_SAMPLE_VALUES]),
            )
        )
    grain_candidates = tuple(
        column.header
        for column in columns
        if column.non_null_count == data_row_count and column.uniqueness_ratio == 1.0
    )
    return SheetProfile(
        name=worksheet.title,
        index=index,
        state=worksheet.sheet_state,
        used_region=CellRegion(1, max_row, 1, max_column),
        header_row=header_row,
        field_id_row=field_id_row,
        data_start_row=data_start if data_end is not None else None,
        data_end_row=data_end,
        data_row_count=data_row_count,
        columns=tuple(columns),
        grain_candidates=grain_candidates,
    )


def profile_workbook(source: Path | bytes) -> WorkbookProfile:
    content = _read_source(source)
    _validate_archive(content)
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=False)
    except (OSError, ValueError, KeyError, zipfile.BadZipFile) as error:
        raise InvalidWorkbookError("source is not a readable XLSX workbook") from error
    try:
        sheets = tuple(
            _profile_sheet(worksheet, index) for index, worksheet in enumerate(workbook.worksheets)
        )
    finally:
        workbook.close()
    return WorkbookProfile(
        sha256=hashlib.sha256(content).hexdigest(),
        size_bytes=len(content),
        sheets=sheets,
    )
