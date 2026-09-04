"""Task 2 契约测试：治理化的 FLOW Excel 模板下载。

被测行为（尚未实现，RED）：
GET /api/v1/intake/templates/flow.excel.v1 返回确定性空白模板：
- 固定安全文件名 + xlsx 内容类型 + no-store；
- 工作表顺序与冻结契约一致；前 3 行为 契约行（显示名/字段ID/填写提示）且受保护；
- 数据区为空；两次请求字节稳定（deterministic renderer）；
- 未知模板 ID 返回 typed 404。
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any

import pytest
from alembic import command
from alembic.config import Config
from httpx import ASGITransport, AsyncClient

from flow_api.data_contract.contract import load_contract
from flow_api.main import create_app

REPOSITORY_ROOT = Path(__file__).resolve().parents[4]
CONTRACT_PATH = REPOSITORY_ROOT / "templates/excel/flow_v1_contract.yaml"

TEMPLATE_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
TEMPLATE_FILENAME = "flow.excel.v1.template.xlsx"


@pytest.fixture(scope="module", autouse=True)
def migrated_database() -> None:
    command.upgrade(Config("alembic.ini"), "head")


@pytest.fixture
async def client() -> Any:
    app = create_app()
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as async_client:
        yield async_client


@pytest.mark.asyncio
async def test_template_download_returns_governed_workbook(client: Any) -> None:
    response = await client.get("/api/v1/intake/templates/flow.excel.v1")
    assert response.status_code == 200
    assert response.headers["content-type"] == TEMPLATE_MIME
    disposition = response.headers["content-disposition"]
    assert f'attachment; filename="{TEMPLATE_FILENAME}"' in disposition
    assert response.headers["cache-control"] == "no-store"
    assert int(response.headers["content-length"]) == len(response.content)


@pytest.mark.asyncio
async def test_template_matches_frozen_contract_layout(client: Any) -> None:
    from openpyxl import load_workbook

    response = await client.get("/api/v1/intake/templates/flow.excel.v1")
    workbook = load_workbook(BytesIO(response.content))
    contract = load_contract(CONTRACT_PATH)

    expected_sheets = [sheet.sheet_name for sheet in contract.sheets]
    assert workbook.sheetnames == ["说明", *expected_sheets]

    instructions = workbook["说明"]
    assert instructions["A1"].value and "flow.excel.v1" in str(instructions["A1"].value)

    for sheet in contract.sheets:
        worksheet = workbook[sheet.sheet_name]
        assert [cell.value for cell in worksheet[1]] == [
            field.display_name for field in sheet.fields
        ]
        assert [cell.value for cell in worksheet[2]] == [field.field_id for field in sheet.fields]
        # 契约行受保护：冻结前三行，数据区从第 4 行开始且为空
        assert worksheet.freeze_panes == "A4"
        assert worksheet.max_row <= 3


@pytest.mark.asyncio
async def test_template_is_byte_stable_across_requests(client: Any) -> None:
    first = (await client.get("/api/v1/intake/templates/flow.excel.v1")).content
    second = (await client.get("/api/v1/intake/templates/flow.excel.v1")).content
    assert first == second


@pytest.mark.asyncio
async def test_unknown_template_returns_typed_404(client: Any) -> None:
    response = await client.get("/api/v1/intake/templates/unknown.template")
    assert response.status_code == 404
    body = response.json()
    assert body["detail"]["code"] == "template_not_found"
