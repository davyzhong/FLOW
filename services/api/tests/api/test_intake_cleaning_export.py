"""Task 4 契约测试：清洗摘要与标准化工作簿导出。

- GET /api/v1/intake/imports/{id}/cleaning-summary：raw/transformed 计数、
  转换规则与版本、有界 before/after 样本（绑定 sheet/row/column 血缘）、
  质量与对账计数；
- GET /api/v1/intake/imports/{id}/standardized-workbook：由 canonical 包确定性
  渲染的 xlsx 导出（派生数据，绝不替代源字节）。
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any

import pytest
from alembic import command
from alembic.config import Config
from botocore.exceptions import ClientError
from httpx import ASGITransport, AsyncClient
from integration.intake_service_support import clean
from sqlalchemy.orm import Session

from flow_api.api.routes.intake import get_db_session, get_source_storage
from flow_api.data_contract.contract import load_contract
from flow_api.infrastructure.db import get_engine
from flow_api.infrastructure.object_store import ObjectStore
from flow_api.intake.source_storage import SourceStorage
from flow_api.main import create_app

REPOSITORY_ROOT = Path(__file__).resolve().parents[4]
NONSTANDARD = REPOSITORY_ROOT / "fixtures/workbooks/external_logistics_nonstandard_v1.xlsx"
CONTRACT_PATH = REPOSITORY_ROOT / "templates/excel/flow_v1_contract.yaml"
WORKBOOK_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class FakeS3Client:
    def __init__(self) -> None:
        self.objects: dict[str, tuple[bytes, str, dict[str, str]]] = {}

    def head_object(self, *, Bucket: str, Key: str) -> dict[str, Any]:  # noqa: N803
        del Bucket
        if Key not in self.objects:
            raise ClientError(
                {
                    "Error": {"Code": "NoSuchKey", "Message": "missing"},
                    "ResponseMetadata": {"HTTPStatusCode": 404},
                },
                "HeadObject",
            )
        content, content_type, metadata = self.objects[Key]
        return {"ContentLength": len(content), "ContentType": content_type, "Metadata": metadata}

    def put_object(
        self,
        *,
        Bucket: str,  # noqa: N803
        Key: str,  # noqa: N803
        Body: bytes,  # noqa: N803
        ContentType: str,  # noqa: N803
        Metadata: dict[str, str],  # noqa: N803
    ) -> None:
        del Bucket
        self.objects[Key] = (Body, ContentType, Metadata)

    def get_object(self, *, Bucket: str, Key: str) -> dict[str, Any]:  # noqa: N803
        del Bucket
        if Key not in self.objects:
            raise ClientError(
                {
                    "Error": {"Code": "NoSuchKey", "Message": "missing"},
                    "ResponseMetadata": {"HTTPStatusCode": 404},
                },
                "HeadObject",
            )
        return {"Body": BytesIO(self.objects[Key][0])}


@pytest.fixture(scope="module", autouse=True)
def migrated_database() -> None:
    command.upgrade(Config("alembic.ini"), "head")


@pytest.fixture
async def ready_import(client: Any) -> dict[str, Any]:
    """走完 upload→map→confirm→validate，得到一个 ready 的导入版本。"""
    batch = (
        await client.post("/api/v1/intake/batches", json={"name": "cleaning summary"})
    ).json()
    upload = await client.post(
        f"/api/v1/intake/batches/{batch['id']}/sources",
        files={"workbook": (NONSTANDARD.name, NONSTANDARD.read_bytes(), WORKBOOK_MIME)},
    )
    source = upload.json()
    mapping = (
        await client.post(f"/api/v1/intake/sources/{source['id']}/mapping-proposals")
    ).json()
    await client.post(
        f"/api/v1/intake/mappings/{mapping['id']}/confirm",
        json={"actor": "finance.bp@example.com"},
    )
    version = (
        await client.post(
            f"/api/v1/intake/sources/{source['id']}/validate",
            json={"mapping_version_id": mapping["id"]},
        )
    ).json()
    return {"source": source, "mapping": mapping, "version": version}


@pytest.fixture
async def client() -> Any:
    with Session(get_engine(), expire_on_commit=False) as session:
        clean(session)
        storage = SourceStorage(ObjectStore(client=FakeS3Client(), bucket="flow"))
        app = create_app()

        def session_override() -> Any:
            yield session
            session.flush()

        app.dependency_overrides[get_db_session] = session_override
        app.dependency_overrides[get_source_storage] = lambda: storage
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as async_client:
            yield async_client
        session.rollback()
        clean(session)


async def test_cleaning_summary_exposes_bounded_lineage(client: Any, ready_import: dict) -> None:
    version_id = ready_import["version"]["id"]
    response = await client.get(f"/api/v1/intake/imports/{version_id}/cleaning-summary")
    assert response.status_code == 200, response.text
    summary = response.json()

    assert summary["import_version_id"] == version_id
    assert summary["totals"]["raw_values"] > 0
    assert summary["totals"]["records"] > 0

    assert summary["transform_rules"], "至少应有一条转换规则"
    for rule in summary["transform_rules"]:
        assert rule["rule_id"] and rule["rule_version"] > 0
        assert rule["applied_count"] > 0
        assert len(rule["samples"]) <= 3
        for sample in rule["samples"]:
            assert {"sheet_name", "source_row", "source_column", "canonical_field"} <= set(sample)
            assert "raw_value" in sample and "transformed_value" in sample

    assert {"blocking", "warning"} <= set(summary["quality_issues"])
    assert {"passed", "failed"} <= set(summary["reconciliation"])


async def test_standardized_workbook_export_is_deterministic(
    client: Any, ready_import: dict
) -> None:
    from openpyxl import load_workbook

    version_id = ready_import["version"]["id"]
    first = await client.get(f"/api/v1/intake/imports/{version_id}/standardized-workbook")
    assert first.status_code == 200, first.text
    assert first.headers["content-type"] == WORKBOOK_MIME
    assert "attachment" in first.headers["content-disposition"]
    assert first.headers["cache-control"] == "no-store"

    second = await client.get(f"/api/v1/intake/imports/{version_id}/standardized-workbook")
    assert first.content == second.content

    contract = load_contract(CONTRACT_PATH)
    workbook = load_workbook(BytesIO(first.content))
    assert workbook.sheetnames == [sheet.sheet_name for sheet in contract.sheets]
    # 标准化导出来自 canonical 数据：契约工作表内应存在数据行
    data_sheet = workbook[contract.sheets[0].sheet_name]
    assert data_sheet.max_row > 3


async def test_summary_missing_import_returns_404(client: Any) -> None:
    from uuid import uuid4

    response = await client.get(f"/api/v1/intake/imports/{uuid4()}/cleaning-summary")
    assert response.status_code == 404
