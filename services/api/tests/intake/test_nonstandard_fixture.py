from __future__ import annotations

import hashlib
import json
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

REPOSITORY_ROOT = Path(__file__).resolve().parents[4]
WORKBOOK_PATH = REPOSITORY_ROOT / "fixtures/workbooks/external_logistics_nonstandard_v1.xlsx"
MANIFEST_PATH = REPOSITORY_ROOT / "fixtures/intake/nonstandard_manifest.json"
GENERATOR = REPOSITORY_ROOT / "scripts/generate_phase_3_fixtures.py"


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def test_nonstandard_fixture_has_intentional_external_structure() -> None:
    manifest = json.loads(MANIFEST_PATH.read_text(encoding="utf-8"))
    workbook = load_workbook(WORKBOOK_PATH, read_only=False, data_only=True)

    assert manifest["fixture_id"] == "external.logistics.nonstandard.v1"
    assert manifest["expected_contract_version"] == "flow.excel.v1"
    assert manifest["workbook_sha256"] == _sha256(WORKBOOK_PATH)
    assert tuple(workbook.sheetnames) == tuple(manifest["sheet_order"])
    assert len(workbook.sheetnames) == 9
    assert not set(workbook.sheetnames) & {
        "01_分析批次",
        "02_经营实际",
        "03_财务实际",
        "04_月度预算",
        "05_应收回款",
        "06_客户主数据",
        "07_物流产品",
        "08_组织与区域",
        "09_管理科目",
    }

    for sheet_name, sheet_manifest in manifest["sheets"].items():
        worksheet = workbook[sheet_name]
        header_row = sheet_manifest["header_row"]
        headers = tuple(cell.value for cell in worksheet[header_row] if cell.value is not None)
        assert header_row == 3
        assert worksheet["A1"].value == "FLOW 外部物流经营数据样例"
        assert worksheet["A2"].value
        assert not set(sheet_manifest["field_aliases"]) & set(headers)
        assert set(sheet_manifest["field_aliases"].values()) <= set(headers)
        assert "数据备注" in headers


def test_nonstandard_fixture_preserves_reference_row_counts_and_variations() -> None:
    manifest = json.loads(MANIFEST_PATH.read_text(encoding="utf-8"))
    workbook = load_workbook(WORKBOOK_PATH, read_only=False, data_only=True)

    for sheet_name, sheet_manifest in manifest["sheets"].items():
        worksheet = workbook[sheet_name]
        assert worksheet.max_row - sheet_manifest["header_row"] == sheet_manifest["row_count"]

    operating = workbook[manifest["roles"]["operating_actual"]]
    headers = {cell.value: cell.column for cell in operating[3]}
    month_values = {
        operating.cell(row=row, column=headers["业务月份"]).value
        for row in range(4, operating.max_row + 1)
    }
    revenue_values = [
        operating.cell(row=row, column=headers["营业收入(元)"]).value
        for row in range(4, min(operating.max_row, 80) + 1)
    ]
    assert any(isinstance(value, str) and "/" in value for value in month_values)
    assert any(isinstance(value, str) and "年" in value for value in month_values)
    assert any(isinstance(value, str) and "," in value for value in revenue_values)
    assert all(cell.value not in {"record_id", "month_key", "revenue"} for cell in operating[2])


def test_nonstandard_fixture_generation_is_byte_deterministic() -> None:
    before = (_sha256(WORKBOOK_PATH), _sha256(MANIFEST_PATH))
    subprocess.run([sys.executable, str(GENERATOR)], cwd=REPOSITORY_ROOT, check=True)
    first = (_sha256(WORKBOOK_PATH), _sha256(MANIFEST_PATH))
    subprocess.run([sys.executable, str(GENERATOR)], cwd=REPOSITORY_ROOT, check=True)
    second = (_sha256(WORKBOOK_PATH), _sha256(MANIFEST_PATH))

    assert first == before
    assert second == first
