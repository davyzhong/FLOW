from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from flow_api.data_contract.contract import load_contract
from flow_api.data_contract.parser import WorkbookParseError, parse_workbook
from flow_api.data_contract.semantic import compare_semantics
from flow_api.fixtures.generator import build_reference_package

REPOSITORY_ROOT = Path(__file__).resolve().parents[4]
CONTRACT = load_contract(REPOSITORY_ROOT / "templates/excel/flow_v1_contract.yaml")
STANDARD_WORKBOOK = REPOSITORY_ROOT / "fixtures/workbooks/flow_standard_v1.xlsx"


def test_standard_workbook_parses_to_reference_package() -> None:
    parsed = parse_workbook(STANDARD_WORKBOOK, CONTRACT)

    assert compare_semantics(build_reference_package(), parsed) == ()


def test_parser_uses_field_ids_not_positions_or_display_names(tmp_path: Path) -> None:
    destination = tmp_path / "reordered.xlsx"
    workbook = load_workbook(STANDARD_WORKBOOK)
    worksheet = workbook["02_经营实际"]
    worksheet["A1"] = "自定义月份标题"
    worksheet["B1"] = "自定义记录标题"
    for row_index in range(1, worksheet.max_row + 1):
        first = worksheet.cell(row=row_index, column=1).value
        second = worksheet.cell(row=row_index, column=2).value
        worksheet.cell(row=row_index, column=1).value = second
        worksheet.cell(row=row_index, column=2).value = first
    workbook.save(destination)

    parsed = parse_workbook(destination, CONTRACT)

    assert compare_semantics(build_reference_package(), parsed) == ()


@pytest.mark.parametrize(
    ("mutation", "expected_code"),
    (
        ("missing_field_id", "missing_field_id"),
        ("duplicate_field_id", "duplicate_field_id"),
        ("unknown_field_id", "unknown_field_id"),
        ("wrong_contract_version", "incompatible_contract_version"),
        ("broken_relationship", "broken_relationship"),
        ("duplicate_grain", "duplicate_grain"),
        ("literal_null", "invalid_type"),
    ),
)
def test_parser_reports_typed_blocking_issues(
    tmp_path: Path, mutation: str, expected_code: str
) -> None:
    destination = tmp_path / f"{mutation}.xlsx"
    workbook = load_workbook(STANDARD_WORKBOOK)

    if mutation == "missing_field_id":
        workbook["02_经营实际"]["I2"] = None
    elif mutation == "duplicate_field_id":
        workbook["02_经营实际"]["I2"] = "order_count"
    elif mutation == "unknown_field_id":
        workbook["02_经营实际"]["I2"] = "unknown_amount"
    elif mutation == "wrong_contract_version":
        workbook["01_分析批次"]["B4"] = "flow.excel.v99"
    elif mutation == "broken_relationship":
        workbook["02_经营实际"]["D4"] = "MISSING_CUSTOMER"
    elif mutation == "duplicate_grain":
        for column_index in range(1, 13):
            workbook["02_经营实际"].cell(row=5, column=column_index).value = (
                workbook["02_经营实际"].cell(row=4, column=column_index).value
            )
    elif mutation == "literal_null":
        workbook["04_月度预算"]["I4"] = "NULL"
    workbook.save(destination)

    with pytest.raises(WorkbookParseError) as error:
        parse_workbook(destination, CONTRACT)

    assert expected_code in {issue.code for issue in error.value.issues}
    assert all(issue.severity == "blocking" for issue in error.value.issues)
