from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from flow_api.data_contract.contract import load_contract
from flow_api.data_contract.workbook import render_workbook, workbook_semantic_fingerprint
from flow_api.fixtures.generator import build_reference_package

REPOSITORY_ROOT = Path(__file__).resolve().parents[4]
CONTRACT_PATH = REPOSITORY_ROOT / "templates/excel/flow_v1_contract.yaml"


def test_generated_workbook_uses_exact_contract_structure(tmp_path: Path) -> None:
    contract = load_contract(CONTRACT_PATH)
    package = build_reference_package()
    destination = tmp_path / "flow.xlsx"

    render_workbook(contract, package, destination)
    workbook = load_workbook(destination, data_only=False)

    assert tuple(workbook.sheetnames) == tuple(sheet.sheet_name for sheet in contract.sheets)
    assert workbook.properties.version == "flow.excel.v1"
    assert workbook.properties.created.isoformat() == "2026-08-30T00:00:00"
    for sheet_contract in contract.sheets:
        worksheet = workbook[sheet_contract.sheet_name]
        assert worksheet.freeze_panes == "A4"
        assert tuple(cell.value for cell in worksheet[1]) == tuple(
            field.display_name for field in sheet_contract.fields
        )
        assert tuple(cell.value for cell in worksheet[2]) == tuple(
            field.field_id for field in sheet_contract.fields
        )
        assert all(cell.comment is not None for cell in worksheet[1])
        assert worksheet.auto_filter.ref is not None


def test_generated_workbook_contains_all_reference_records(tmp_path: Path) -> None:
    contract = load_contract(CONTRACT_PATH)
    package = build_reference_package()
    destination = tmp_path / "flow.xlsx"
    render_workbook(contract, package, destination)
    workbook = load_workbook(destination, read_only=False, data_only=True)

    expected_rows = {
        "00_填写说明": 7,
        "01_分析批次": 1,
        "02_经营实际": len(package.operating_actuals),
        "03_财务实际": len(package.financial_actuals),
        "04_月度预算": len(package.monthly_budgets),
        "05_应收回款": len(package.ar_collections),
        "06_客户主数据": len(package.customers),
        "07_物流产品": len(package.logistics_products),
        "08_组织与区域": len(package.organizations) + len(package.regions),
        "09_管理科目": len(package.management_accounts),
    }
    for sheet_name, count in expected_rows.items():
        assert workbook[sheet_name].max_row == count + 3

    batch_sheet = workbook["01_分析批次"]
    batch_values = dict(
        zip(
            (cell.value for cell in batch_sheet[2]),
            (cell.value for cell in batch_sheet[4]),
            strict=True,
        )
    )
    assert batch_values["contract_version"] == "flow.excel.v1"
    assert batch_values["analysis_start_month"] == "2025-09"
    assert batch_values["analysis_end_month"] == "2026-08"


def test_workbook_is_safe_fillable_and_has_validations(tmp_path: Path) -> None:
    contract = load_contract(CONTRACT_PATH)
    destination = tmp_path / "flow.xlsx"
    render_workbook(contract, build_reference_package(), destination)
    workbook = load_workbook(destination, data_only=False)

    assert not workbook.vba_archive
    assert not workbook._external_links  # noqa: SLF001
    assert all(worksheet.sheet_state == "visible" for worksheet in workbook.worksheets)
    for worksheet in workbook.worksheets:
        assert not any(
            cell.data_type == "f"
            for row in worksheet.iter_rows()
            for cell in row
            if cell.value is not None
        )
        assert worksheet.protection.sheet
        assert all(cell.protection.locked for cell in worksheet[2])
        if worksheet.max_row >= 4:
            assert all(not cell.protection.locked for cell in worksheet[4])

    assert len(workbook["08_组织与区域"].data_validations.dataValidation) >= 1
    assert len(workbook["09_管理科目"].data_validations.dataValidation) >= 1
    instructions = " ".join(
        str(cell.value)
        for row in workbook["00_填写说明"].iter_rows(min_row=4)
        for cell in row
        if cell.value
    )
    assert "直接填写" in instructions
    assert "非标准" in instructions
    assert "字段 ID" in instructions
    assert workbook["00_填写说明"].column_dimensions["C"].width >= 70
    assert workbook["00_填写说明"]["C4"].alignment.wrap_text


def test_workbook_semantic_fingerprint_is_stable(tmp_path: Path) -> None:
    contract = load_contract(CONTRACT_PATH)
    package = build_reference_package()
    first = tmp_path / "first.xlsx"
    second = tmp_path / "second.xlsx"

    render_workbook(contract, package, first)
    render_workbook(contract, package, second)

    assert workbook_semantic_fingerprint(first) == workbook_semantic_fingerprint(second)
